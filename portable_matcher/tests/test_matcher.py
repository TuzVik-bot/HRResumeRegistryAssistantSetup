from pathlib import Path

from openpyxl import Workbook, load_workbook

from hr_matcher.app import run_match
from hr_matcher.excel_io import read_candidates
from hr_matcher.matcher import match_candidates, scan_resumes


def test_matches_russian_registry_to_latin_filename(tmp_path):
    candidate_file = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["номер", "Фамилия, имя \n(на русском языке)", "вакансия"])
    sheet.append([4371, "Громов Сергей", "программист embedded"])
    sheet.append([4372, "Петрова Ольга", "программист embedded"])
    workbook.save(candidate_file)

    resumes_dir = tmp_path / "resumes"
    resumes_dir.mkdir()
    (resumes_dir / "CV_Siarhei_Hromau_Embedded_engineer.pdf").write_text("x", encoding="utf-8")
    (resumes_dir / "Olga_Petrova_CV.pdf").write_text("x", encoding="utf-8")

    candidates = read_candidates(candidate_file)
    resumes = scan_resumes(resumes_dir)
    results = match_candidates(candidates, resumes)

    assert [result.status for result in results] == ["matched", "matched"]
    assert results[0].resume.original_filename == "CV_Siarhei_Hromau_Embedded_engineer.pdf"
    assert results[1].resume.original_filename == "Olga_Petrova_CV.pdf"


def test_run_match_writes_copies_and_excel_report(tmp_path):
    registry = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["id", "Фамилия, имя", "вакансия"])
    sheet.append([10, "Иванова Анна Сергеевна", "BA"])
    workbook.save(registry)

    resumes = tmp_path / "resumes"
    resumes.mkdir()
    (resumes / "Иванова Анна Сергеевна (2).pdf").write_text("x", encoding="utf-8")
    output = tmp_path / "out"

    summary = run_match(registry, resumes, output)

    assert summary.matched == 1
    assert summary.report_path.exists()
    assert list((output / "matched_resumes").glob("*.pdf"))
    report = load_workbook(summary.report_path, read_only=True)
    assert "Все результаты" in report.sheetnames


def test_unknown_resume_goes_to_unmatched(tmp_path):
    registry = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["ФИО", "вакансия"])
    sheet.append(["Семенова Ольга", "руководитель проектов"])
    workbook.save(registry)
    resumes = tmp_path / "resumes"
    resumes.mkdir()
    (resumes / "CV_Siarhei_Hromau.pdf").write_text("x", encoding="utf-8")

    results = match_candidates(read_candidates(registry), scan_resumes(resumes))

    assert results[0].status == "unmatched"
    assert results[0].resume is None


def test_technical_registry_uses_generated_candidate_code(tmp_path):
    registry = tmp_path / "technical.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["first_name", "last_name", "source", "file_name", "vacancy", "stage"])
    sheet.append(["Имя", "Фамилия", "Источник", "Файл", "Вакансия", "Этап"])
    sheet.append(["Фамилия, имя", "", "", "", "вакансия", "статус"])
    sheet.append(["Иван", "Иванов", "rabota.by", "", "BA", "в работе"])
    workbook.save(registry)

    candidates = read_candidates(registry)

    assert candidates[0].candidate_code == "CAND-000004"
    assert candidates[0].full_name == "Иванов Иван"


def test_matches_when_resume_omits_patronymic_and_uses_latin_order(tmp_path):
    registry = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["id", "Фамилия, имя", "вакансия"])
    sheet.append([1, "Иванов Иван Иванович", "BA"])
    workbook.save(registry)
    resumes = tmp_path / "resumes"
    resumes.mkdir()
    (resumes / "Ivan_Ivanov_resume.pdf").write_text("x", encoding="utf-8")

    results = match_candidates(read_candidates(registry), scan_resumes(resumes))

    assert results[0].status == "matched"
    assert results[0].resume.original_filename == "Ivan_Ivanov_resume.pdf"


def test_matches_compact_filename_without_spaces(tmp_path):
    registry = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["id", "Фамилия, имя", "вакансия"])
    sheet.append([1, "Иванов Иван", "BA"])
    workbook.save(registry)
    resumes = tmp_path / "resumes"
    resumes.mkdir()
    (resumes / "IvanovIvan.pdf").write_text("x", encoding="utf-8")

    results = match_candidates(read_candidates(registry), scan_resumes(resumes))

    assert results[0].status == "matched"


def test_first_name_only_does_not_create_false_review(tmp_path):
    registry = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(["ФИО", "вакансия"])
    sheet.append(["Виталий Романовский", "BA"])
    workbook.save(registry)
    resumes = tmp_path / "resumes"
    resumes.mkdir()
    (resumes / "Siarhei_Hromau_CV.pdf").write_text("x", encoding="utf-8")

    results = match_candidates(read_candidates(registry), scan_resumes(resumes))

    assert results[0].status == "unmatched"
    assert results[0].resume is None
