from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from hr_matcher.models import MatchResult, ResumeAudit, ResumeFile


def write_report(
    results: list[MatchResult],
    resumes: list[ResumeFile],
    output_dir: Path,
    resume_audits: list[ResumeAudit] | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    ws_all = workbook.active
    ws_all.title = "Все результаты"
    _write_results_sheet(ws_all, results)
    _write_results_sheet(workbook.create_sheet("Найдено"), [result for result in results if result.status == "matched"])
    _write_results_sheet(workbook.create_sheet("На проверку"), [result for result in results if result.status == "review"])
    _write_results_sheet(workbook.create_sheet("Не найдено"), [result for result in results if result.status == "unmatched"])
    matched_resume_paths = {result.resume.path for result in results if result.resume}
    unassigned_resumes = [resume for resume in resumes if resume.path not in matched_resume_paths]
    _write_unmatched_resumes(workbook.create_sheet("Лишние резюме"), unassigned_resumes, resume_audits or [])
    if resume_audits is not None:
        _write_resume_audit(workbook.create_sheet("Проверка резюме"), resume_audits)
    report_path = output_dir / "resume_match_report.xlsx"
    workbook.save(report_path)
    return report_path


def _write_results_sheet(sheet, results: list[MatchResult]) -> None:
    source_headers = _source_headers(results)
    headers = [
        "Статус",
        "Score",
        "Второй score",
        "ID/номер",
        "Строка Excel",
        "ФИО из Excel",
        "Вакансия",
        "Файл резюме",
        "Скопирован как",
        "Причина",
    ] + source_headers
    sheet.append(headers)
    _style_header(sheet)
    for result in results:
        sheet.append(
            [
                _status_label(result.status),
                result.score,
                result.second_score,
                result.candidate.candidate_code,
                result.candidate.row_number,
                result.candidate.full_name,
                result.candidate.vacancy,
                result.resume.original_filename if result.resume else "",
                str(result.output_path) if result.output_path else "",
                result.reason,
            ]
            + [result.candidate.row_values.get(header, "") for header in source_headers]
        )
    _autosize(sheet)


def _write_unmatched_resumes(sheet, resumes: list[ResumeFile], audits: list[ResumeAudit]) -> None:
    audit_by_path = {audit.resume.path: audit for audit in audits}
    sheet.append(["Файл резюме", "Путь", "Лучший кандидат", "Строка Excel", "Score", "Статус", "Причина"])
    _style_header(sheet)
    for resume in resumes:
        audit = audit_by_path.get(resume.path)
        candidate = audit.best_candidate if audit else None
        sheet.append(
            [
                resume.original_filename,
                str(resume.path),
                candidate.full_name if candidate else "",
                candidate.row_number if candidate else "",
                audit.score if audit else "",
                _status_label(audit.status) if audit else "",
                audit.reason if audit else "",
            ]
        )
    _autosize(sheet)


def _write_resume_audit(sheet, audits: list[ResumeAudit]) -> None:
    sheet.append(["Закреплено", "Статус", "Score", "Второй score", "Файл резюме", "Лучший кандидат", "Строка Excel", "ID/номер", "Вакансия", "Причина"])
    _style_header(sheet)
    for audit in audits:
        candidate = audit.best_candidate
        sheet.append(
            [
                "да" if audit.assigned else "нет",
                _status_label(audit.status),
                audit.score,
                audit.second_score,
                audit.resume.original_filename,
                candidate.full_name if candidate else "",
                candidate.row_number if candidate else "",
                candidate.candidate_code if candidate else "",
                candidate.vacancy if candidate else "",
                audit.reason,
            ]
        )
    _autosize(sheet)


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 80)


def _status_label(status: str) -> str:
    return {"matched": "найдено", "review": "на проверку", "unmatched": "не найдено"}.get(status, status)


def _source_headers(results: list[MatchResult]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for result in results:
        for header in result.candidate.row_values:
            if header not in seen:
                seen.add(header)
                headers.append(header)
    return headers
