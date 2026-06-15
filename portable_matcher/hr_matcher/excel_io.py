from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from hr_matcher.models import Candidate
from hr_matcher.name_tools import normalize_text


FULL_NAME_ALIASES = ("фио", "фамилия имя", "фамилия имя на русском языке", "кандидат")
VACANCY_ALIASES = ("вакансия", "позиция", "должность", "vacancy")


def read_candidates(workbook_path: Path) -> list[Candidate]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = _main_sheet(workbook)
    header_row, headers = _find_header(sheet)
    full_name_col = _find_column(headers, FULL_NAME_ALIASES)
    vacancy_col = _find_column(headers, VACANCY_ALIASES)
    first_name_col = _find_exact(headers, "first name") or _find_exact(headers, "first_name")
    last_name_col = _find_exact(headers, "last name") or _find_exact(headers, "last_name")
    is_technical = first_name_col is not None and last_name_col is not None
    if full_name_col is None and not is_technical:
        raise ValueError("Не найдена колонка ФИО в Excel-реестре")

    data_start = header_row + (3 if is_technical else 1)
    candidates: list[Candidate] = []
    empty_after_data = 0
    for row_number, values in _iter_rows(sheet, start_row=data_start):
        if not any(_has_value(value) for value in values):
            if candidates:
                empty_after_data += 1
                if empty_after_data >= 200:
                    break
            continue
        empty_after_data = 0
        row_values = _row_dict(headers, values)
        full_name = ""
        if full_name_col is not None:
            full_name = _cell_text(values, full_name_col)
        elif is_technical:
            full_name = " ".join(part for part in [_cell_text(values, last_name_col), _cell_text(values, first_name_col)] if part)
        if len(full_name.split()) < 2:
            continue
        source_id = "" if is_technical or full_name_col == 0 else _cell_text(values, 0)
        candidate_code = source_id if source_id and source_id.lower() not in {"none", "nan"} else f"CAND-{row_number:06d}"
        candidates.append(
            Candidate(
                row_number=row_number,
                candidate_code=candidate_code,
                full_name=full_name,
                vacancy=_cell_text(values, vacancy_col) if vacancy_col is not None else "",
                row_values=row_values,
            )
        )
    return candidates


def _main_sheet(workbook):
    for sheet in workbook.worksheets:
        if sheet.title.strip().lower() == "реестр":
            return sheet
    return workbook.active


def _find_header(sheet) -> tuple[int, list[str]]:
    checked = 0
    for row_number, values in _iter_rows(sheet, start_row=1):
        if not any(_has_value(value) for value in values):
            continue
        headers = [str(value or "").strip() for value in values]
        normalized = [normalize_text(header) for header in headers]
        score = sum(1 for alias in FULL_NAME_ALIASES + VACANCY_ALIASES if any(alias in header for header in normalized))
        score += 2 if "first name" in normalized and "last name" in normalized else 0
        if score:
            return row_number, headers
        checked += 1
        if checked >= 30:
            break
    raise ValueError("Не найдена строка заголовков на листе реестр")


def _iter_rows(sheet, start_row: int):
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        yield row[0].row if hasattr(row[0], "row") else start_row, row
        start_row += 1


def _find_column(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    normalized = [normalize_text(header) for header in headers]
    for alias in aliases:
        for index, header in enumerate(normalized):
            if alias and alias in header:
                return index
    return None


def _find_exact(headers: list[str], name: str) -> int | None:
    normalized = [normalize_text(header).replace(" ", "_") for header in headers]
    for index, header in enumerate(normalized):
        if header == name.replace(" ", "_"):
            return index
    return None


def _row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for index, value in enumerate(values):
        header = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
        result[str(header)] = value
    return result


def _cell_text(values: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    value = values[index]
    return str(value).strip() if value is not None else ""


def _has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""
