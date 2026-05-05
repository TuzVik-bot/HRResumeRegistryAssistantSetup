from openpyxl import Workbook

from app import database
from app.registry import import_registry


def test_imports_technical_registry_and_skips_duplicate(tmp_path, isolated_project_files):
    database.init_db()
    workbook_path = tmp_path / "technical.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "реестр"
    sheet.append(
        [
            "first_name",
            "last_name",
            "source",
            "file_name",
            "resume_link",
            "resume_social",
            "vacancy",
            "stage",
            "reject",
            "recruiter",
            "salary",
            "currency",
            "email",
            "phone",
            "notes",
        ]
    )
    sheet.append(["Имя", "Фамилия", "Источник", "Файл", "PDF", "Профиль", "Вакансия", "Этап", "Отказ", "Рекрутер", "Зарплата", "Валюта", "Email", "Телефон", "Комментарий"])
    sheet.append(["Фамилия, имя", "", "", "", "", "", "вакансия", "статус", "состояние", "ответственный рекрутер", "зарплата", "валюта", "", "", "комментарий"])
    sheet.append(["Иван", "Иванов", "rabota.by", "", "", "", "Business Analyst", "в работе", "", "Довгалёнок", "800-900", "$", "", "", "Structured BA profile"])
    sheet.append(["Мария", "Петрова", "hh.ru", "", "", "", "Business Analyst", "отказ", "не подошла по стеку", "Довгалёнок", "2500", "р", "", "", "Strong communication"])
    workbook.save(workbook_path)

    result = import_registry(workbook_path, "technical.xlsx", file_hash="hash-technical")

    candidates = database.fetch_all("candidates")
    assert result["source_schema"] == "technical"
    assert result["rows"] == 2
    assert len(candidates) == 2
    assert candidates[0]["candidate_id"] == "REG001-CAND-000004"
    assert candidates[0]["full_name"] == "Иванов Иван"
    assert "salary request format mismatch" not in candidates[0]["quality_warnings_json"]

    duplicate = import_registry(workbook_path, "technical.xlsx", file_hash="hash-technical")
    assert duplicate["duplicate"] is True
    assert len(database.fetch_all("candidates")) == 2


def test_accumulates_candidates_from_multiple_registry_sources(tmp_path, isolated_project_files):
    database.init_db()

    legacy_path = tmp_path / "legacy.xlsx"
    legacy = Workbook()
    legacy_sheet = legacy.active
    legacy_sheet.title = "реестр"
    legacy_sheet.append(["Фамилия, имя", "вакансия", "статус", "ответственный рекрутер", "запрос кандидата, указывается р/$", "оценка рекрутера"])
    legacy_sheet.append(["Иванов Иван", "Python developer", "в работе", "Анна", "3000$", "Django"])
    legacy.save(legacy_path)

    technical_path = tmp_path / "technical.xlsx"
    technical = Workbook()
    technical_sheet = technical.active
    technical_sheet.title = "реестр"
    technical_sheet.append(["first_name", "last_name", "source", "vacancy", "stage", "recruiter", "salary", "currency", "notes"])
    technical_sheet.append(["Имя", "Фамилия", "Источник", "Вакансия", "Этап", "Рекрутер", "Зарплата", "Валюта", "Комментарий"])
    technical_sheet.append(["Фамилия, имя", "", "", "вакансия", "статус", "ответственный рекрутер", "зарплата", "валюта", "комментарий"])
    technical_sheet.append(["Мария", "Петрова", "hh.ru", "Business Analyst", "отказ", "Ольга", "2500", "р", "notes"])
    technical.save(technical_path)

    import_registry(legacy_path, "legacy.xlsx", file_hash="hash-legacy")
    import_registry(technical_path, "technical.xlsx", file_hash="hash-technical")

    registries = database.fetch_all("registries")
    candidates = database.fetch_all("candidates")
    assert len(registries) == 2
    assert len(candidates) == 2
    assert {candidate["full_name"] for candidate in candidates} == {"Иванов Иван", "Петрова Мария"}
